"""Export service for CRISPR-Sim: PDF, Excel (XLSX), CSV, FASTA."""

from __future__ import annotations

import io
import csv
from datetime import datetime
from typing import Any, Dict, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable


def generate_analysis_pdf(data: Dict[str, Any]) -> bytes:
    """Generates an enterprise-grade scientific PDF analysis report."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40,
    )

    styles = getSampleStyleSheet()
    primary_color = colors.HexColor("#0D9488")
    dark_bg = colors.HexColor("#1E293B")
    text_dark = colors.HexColor("#0F172A")
    danger_color = colors.HexColor("#DC2626")
    success_color = colors.HexColor("#16A34A")

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=26,
        textColor=primary_color,
    )

    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#64748B"),
    )

    heading_style = ParagraphStyle(
        "SectionHeading",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=13,
        leading=16,
        textColor=dark_bg,
        spaceAfter=6,
    )

    body_style = ParagraphStyle(
        "BodyDark",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=text_dark,
    )

    mono_style = ParagraphStyle(
        "MonoSeq",
        parent=styles["Normal"],
        fontName="Courier",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#1E293B"),
    )

    story = []

    # Title & Metadata Header
    story.append(Paragraph("CRISPR-Sim — Gene Editing Analysis Report", title_style))
    story.append(Paragraph(f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')} | Simulated via CRISPR-Cas9 Engine", subtitle_style))
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=1.5, color=primary_color, spaceAfter=14))

    # Executive Summary Box
    summary_text = data.get("summary", "CRISPR simulation analysis completed.")
    repair_type = data.get("repair_type", "NHEJ")
    frameshift = data.get("frameshift", False)
    premature_stop = data.get("premature_stop", False)
    safety_score = data.get("safety_score", 62)
    safety_label = data.get("safety_label", "Moderate")

    summary_data = [
        [
            Paragraph("<b>Repair Mechanism:</b>", body_style), Paragraph(str(repair_type), body_style),
            Paragraph("<b>Safety Score:</b>", body_style), Paragraph(f"<b>{safety_score}/100</b> ({safety_label})", body_style),
        ],
        [
            Paragraph("<b>Frameshift Mutation:</b>", body_style),
            Paragraph(f"<font color='{'#DC2626' if frameshift else '#16A34A'}'><b>{'YES (Reading Frame Disrupted)' if frameshift else 'NO (In-Frame)'}</b></font>", body_style),
            Paragraph("<b>Premature Stop Codon:</b>", body_style),
            Paragraph(f"<font color='{'#DC2626' if premature_stop else '#16A34A'}'><b>{'YES (Truncated Protein)' if premature_stop else 'NO'}</b></font>", body_style),
        ],
    ]
    summary_table = Table(summary_data, colWidths=[120, 140, 110, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor("#E2E8F0")),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 14))

    # Section 1: Sequence Statistics
    story.append(Paragraph("1. Sequence Statistics", heading_style))
    orig_len = data.get("original_length", 276)
    edit_len = data.get("edited_length", 269)
    len_diff = data.get("length_diff", 7)
    indel_type = "deletion" if len_diff > 0 else "insertion" if len_diff < 0 else "unchanged"

    stats_data = [
        [Paragraph("<b>Metric</b>", body_style), Paragraph("<b>Value</b>", body_style), Paragraph("<b>Clinical/Biological Impact</b>", body_style)],
        [Paragraph("Original DNA Length", body_style), Paragraph(f"{orig_len} bp", body_style), Paragraph("Wild-type sequence baseline", body_style)],
        [Paragraph("Edited DNA Length", body_style), Paragraph(f"{edit_len} bp", body_style), Paragraph("Post-repair sequence length", body_style)],
        [Paragraph("Length Difference (Indel)", body_style), Paragraph(f"{abs(len_diff)} bp ({indel_type})", body_style), Paragraph(f"{'Induces downstream frameshift' if frameshift else 'Preserves codon triplet reading frame'}", body_style)],
        [Paragraph("Predicted Repair Type", body_style), Paragraph(str(repair_type), body_style), Paragraph("Non-Homologous End Joining / HDR", body_style)],
    ]
    stats_table = Table(stats_data, colWidths=[150, 130, 240])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0D9488")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 14))

    # Section 2: Protein Translation Comparison
    story.append(Paragraph("2. Protein Translation Comparison", heading_style))
    orig_protein = data.get("original_protein", "MVHLTPEEKSAVTALWGKVNVDEVGGEALGRLLVVYPWTQRFFE...")
    edit_protein = data.get("edited_protein", "MVHLTPEEKSAVTAPR*TWMKLVVRPWAGCWVVSTLGPRGSLSS...")

    protein_data = [
        [Paragraph("<b>Type</b>", body_style), Paragraph("<b>Amino Acid Sequence (N → C)</b>", body_style)],
        [Paragraph("<b>Original Protein:</b>", body_style), Paragraph(orig_protein, mono_style)],
        [Paragraph("<b>Edited Protein:</b>", body_style), Paragraph(f"<font color='#DC2626'>{edit_protein}</font>", mono_style)],
    ]
    protein_table = Table(protein_data, colWidths=[110, 410])
    protein_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E293B")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(protein_table)
    story.append(Spacer(1, 14))

    # Section 3: mRNA Comparison
    story.append(Paragraph("3. mRNA Sequence Comparison", heading_style))
    orig_mrna = data.get("original_mrna", "AUGGUGCACCUGACUCCUGAGGAGAAGUCUGCCGUUACUGCCCUGUGGGGCAAGGUGAACGUGGAUGAAGUUGGUGGUGAG...")
    edit_mrna = data.get("edited_mrna", "AUGGUGCACCUGACUCCUGAGGAGAAGUCUGCCGUUACUGCCCUGUGGGUGAGAGGCCCUGUGGGGCAAGGUGAACGUGGA...")

    mrna_data = [
        [Paragraph("<b>Original mRNA (5'→3'):</b>", body_style), Paragraph(orig_mrna[:180] + ("..." if len(orig_mrna) > 180 else ""), mono_style)],
        [Paragraph("<b>Edited mRNA (5'→3'):</b>", body_style), Paragraph(edit_mrna[:180] + ("..." if len(edit_mrna) > 180 else ""), mono_style)],
    ]
    mrna_table = Table(mrna_data, colWidths=[130, 390])
    mrna_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(mrna_table)
    story.append(Spacer(1, 14))

    # Footer note
    story.append(Paragraph("<i>This report is generated by CRISPR-Sim for research and academic validation purposes.</i>", subtitle_style))

    doc.build(story)
    return buffer.getvalue()


def generate_analysis_excel(data: Dict[str, Any]) -> bytes:
    """Generates an Excel workbook with multiple structured sheets."""
    wb = openpyxl.Workbook()
    
    # Fonts & Fills
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True, color="0F172A")
    bold_font = Font(name="Arial", size=10, bold=True)
    mono_font = Font(name="Consolas", size=10)
    regular_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # ── Sheet 1: Executive Summary ──
    ws1 = wb.active
    ws1.title = "Summary & Metrics"
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:D1")
    ws1["A1"] = "CRISPR-Sim — Gene Editing Analysis Report"
    ws1["A1"].font = title_font

    ws1["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
    ws1["A2"].font = Font(name="Arial", size=9, italic=True, color="64748B")

    summary_rows = [
        ("Parameter", "Value", "Notes"),
        ("Repair Type", data.get("repair_type", "NHEJ"), "Simulated DNA double-strand break repair"),
        ("Safety Score", f"{data.get('safety_score', 62)}/100", data.get("safety_label", "Moderate")),
        ("Frameshift Mutation", "YES" if data.get("frameshift") else "NO", "Indel disrupts reading frame" if data.get("frameshift") else "Frame preserved"),
        ("Premature Stop Codon", "YES" if data.get("premature_stop") else "NO", "Stop codon truncates protein" if data.get("premature_stop") else "Full protein intact"),
        ("Original DNA Length", f"{data.get('original_length', 276)} bp", "Target sequence"),
        ("Edited DNA Length", f"{data.get('edited_length', 269)} bp", "Post-cleavage sequence"),
        ("Net Indel Size", f"{abs(data.get('length_diff', 7))} bp", "7 bp deletion"),
    ]

    for r_idx, row in enumerate(summary_rows, start=4):
        for c_idx, val in enumerate(row, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if r_idx == 4:
                cell.fill = header_fill
                cell.font = header_font
            elif c_idx == 1:
                cell.font = bold_font
            else:
                cell.font = regular_font

    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 35

    # ── Sheet 2: Protein Comparison ──
    ws2 = wb.create_sheet(title="Protein Comparison")
    ws2.views.sheetView[0].showGridLines = True

    ws2.append(["Type", "Sequence"])
    ws2.append(["Original Protein", data.get("original_protein", "")])
    ws2.append(["Edited Protein", data.get("edited_protein", "")])

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in ws2.iter_rows(min_row=2, max_row=3, min_col=1, max_col=2):
        row[0].font = bold_font
        row[1].font = mono_font

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 100

    # ── Sheet 3: mRNA Comparison ──
    ws3 = wb.create_sheet(title="mRNA Sequences")
    ws3.views.sheetView[0].showGridLines = True

    ws3.append(["Type", "Sequence"])
    ws3.append(["Original mRNA", data.get("original_mrna", "")])
    ws3.append(["Edited mRNA", data.get("edited_mrna", "")])

    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in ws3.iter_rows(min_row=2, max_row=3, min_col=1, max_col=2):
        row[0].font = bold_font
        row[1].font = mono_font

    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 100

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_analysis_csv(data: Dict[str, Any]) -> str:
    """Generates standard CSV output of the analysis results."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Metric", "Value"])
    writer.writerow(["Generated At", datetime.utcnow().isoformat()])
    writer.writerow(["Repair Type", data.get("repair_type", "NHEJ")])
    writer.writerow(["Safety Score", data.get("safety_score", 62)])
    writer.writerow(["Safety Label", data.get("safety_label", "Moderate")])
    writer.writerow(["Frameshift", data.get("frameshift", False)])
    writer.writerow(["Premature Stop Codon", data.get("premature_stop", False)])
    writer.writerow(["Original Length (bp)", data.get("original_length", 276)])
    writer.writerow(["Edited Length (bp)", data.get("edited_length", 269)])
    writer.writerow(["Length Difference (bp)", data.get("length_diff", 7)])
    writer.writerow(["Original Protein", data.get("original_protein", "")])
    writer.writerow(["Edited Protein", data.get("edited_protein", "")])
    writer.writerow(["Original mRNA", data.get("original_mrna", "")])
    writer.writerow(["Edited mRNA", data.get("edited_mrna", "")])
    return output.getvalue()


def generate_analysis_fasta(data: Dict[str, Any]) -> str:
    """Generates FASTA format with headers for original and edited sequences."""
    repair_type = data.get("repair_type", "NHEJ")
    orig_seq = data.get("original_dna", data.get("original_mrna", ""))
    edit_seq = data.get("edited_dna", data.get("edited_mrna", ""))
    orig_prot = data.get("original_protein", "")
    edit_prot = data.get("edited_protein", "")

    fasta_lines = [
        f">CRISPR_Sim|Original_DNA|Length={data.get('original_length', len(orig_seq))}",
        orig_seq,
        f">CRISPR_Sim|Edited_DNA_{repair_type}|Length={data.get('edited_length', len(edit_seq))}|Frameshift={data.get('frameshift', False)}",
        edit_seq,
        f">CRISPR_Sim|Original_Protein|Length={len(orig_prot)}",
        orig_prot,
        f">CRISPR_Sim|Edited_Protein_{repair_type}|Length={len(edit_prot)}|StopCodon={data.get('premature_stop', False)}",
        edit_prot,
    ]
    return "\n".join(fasta_lines)
